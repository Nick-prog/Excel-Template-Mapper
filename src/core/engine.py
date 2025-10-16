from typing import Dict, List, Any, Optional, Union, Tuple
from datetime import datetime, date, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell, MergedCell
from .models import MappingSpec, SheetMapping, ColumnMapping, TRANSFORM_CHOICES
from .utils import is_blank, safe_str, read_workbook_headers, suggest_header_mapping

LONG_DATE_FORMATS: Tuple[str, ...] = (
    "%Y-%m-%d",
    "%m/%d/%Y",
    "%m/%d/%y",
    "%d/%m/%Y",
    "%m-%d-%Y",
    "%Y%m%d",
    "%B %d, %Y",
    "%B %d %Y",
    "%b %d, %Y",
    "%b %d %Y",
    "%d %B %Y",
    "%d %b %Y",
)

def build_initial_spec(template_path: str, source_path: Optional[str] = None) -> MappingSpec:
    template_headers: Dict[str, List[str]] = read_workbook_headers(template_path)
    source_headers: Dict[str, List[str]] = read_workbook_headers(source_path) if source_path else {}
    spec: MappingSpec = MappingSpec(template_path=template_path, source_path=source_path)

    for sheet_name, tgt_headers in template_headers.items():
        sm: SheetMapping = SheetMapping(
            target_sheet=sheet_name,
            target_headers=list(tgt_headers),
            source_sheet=None,
            columns=[ColumnMapping(target=h) for h in tgt_headers],
        )
        if source_path and source_headers:
            best_sheet: Optional[str] = None
            best_score: int = -1
            for sname, sheaders in source_headers.items():
                overlap: int = len(set([h.lower() for h in tgt_headers]) & set([h.lower() for h in sheaders]))
                if overlap > best_score:
                    best_score = overlap
                    best_sheet = sname
            sm.source_sheet = best_sheet
            if best_sheet:
                suggested: Dict[str, str] = suggest_header_mapping(tgt_headers, source_headers.get(best_sheet, []))
                for col in sm.columns:
                    col.source = suggested.get(col.target) or None
        spec.sheets.append(sm)
    return spec

def apply_transforms(value: Any, transforms: List[str]) -> Any:
    v: Any = value
    for t in transforms:
        try:
            if t == "trim":
                v = safe_str(v).strip()
            elif t == "upper":
                v = safe_str(v).upper()
            elif t == "lower":
                v = safe_str(v).lower()
            elif t == "title":
                v = safe_str(v).title()
            elif t == "to_string":
                v = safe_str(v)
            elif t == "to_int":
                if is_blank(v):
                    v = None
                else:
                    if isinstance(v, bool):
                        v = int(v)
                    else:
                        v = int(float(safe_str(v)))
            elif t == "to_float":
                if is_blank(v):
                    v = None
                else:
                    if isinstance(v, bool):
                        v = float(int(v))
                    else:
                        v = float(safe_str(v).replace(",", ""))
            elif t == "date_to_iso":
                if isinstance(v, (datetime, date)):
                    v = v.strftime("%Y-%m-%d")
                else:
                    v = v
            elif t == "digits_only":
                import re
                v = re.sub(r"\D+", "", safe_str(v))
        except Exception:
            pass
    return v

def replace_values(value: Any, repl: Dict[str, str]) -> Any:
    if not repl:
        return value
    sval: str = safe_str(value)
    if sval in repl:
        return repl[sval]
    for k, v in repl.items():
        if k == "":
            continue
        if k in sval:
            sval = sval.replace(k, v)
    return sval

def coerce_value(value: Any, data_type: Optional[str]) -> Any:
    if not data_type or data_type == "general":
        return value
    try:
        if data_type in ("text", "string"):
            return safe_str(value)
        elif data_type in ("integer", "int"):
            if is_blank(value):
                return None
            if isinstance(value, bool):
                return int(value)
            return int(float(safe_str(value).replace(",", "")))
        elif data_type in ("float", "number"):
            if is_blank(value):
                return None
            if isinstance(value, bool):
                return float(int(value))
            return float(safe_str(value).replace(",", ""))
        elif data_type == "boolean":
            s: str = safe_str(value).strip().lower()
            if s in ("y", "yes", "true", "1", "t"):
                return True
            if s in ("n", "no", "false", "0", "f", ""):
                return False
            return bool(value)
        elif data_type == "date":
            if isinstance(value, (datetime, date)):
                return value
            s: str = safe_str(value).strip()
            if s == "":
                return None
            for fmt in LONG_DATE_FORMATS:
                try:
                    return datetime.strptime(s, fmt)
                except Exception:
                    pass
            try:
                serial: float = float(s)
                base: datetime = datetime(1899, 12, 30)
                return base + timedelta(days=serial)
            except Exception:
                return value
        else:
            return value
    except Exception:
        return value

def apply_advanced_to_cell(col: ColumnMapping, row_map: Dict[str, Any], current_value: Any) -> Any:
    # If advanced_format is set, execute it as Python code
    code: Optional[str] = getattr(col, "advanced_format", None)
    if isinstance(code, str) and code.strip():
        # Try to get source dict from local context if available
        import inspect
        frame = inspect.currentframe()
        source_dict: Optional[Dict[str, Any]] = None
        # Look up the call stack for 'source_dict'
        while frame:
            if "source_dict" in frame.f_locals:
                source_dict = frame.f_locals["source_dict"]
                break
            frame = frame.f_back
        local_vars: Dict[str, Any] = {"col": row_map, "source": source_dict, "current_value": current_value}
        try:
            exec(code, {}, local_vars)
            # Expect a function named format_column(col, source)
            if "format_column" in local_vars:
                return local_vars["format_column"](row_map, source_dict)
        except Exception:
            return current_value
    # Otherwise, use rules logic
    rules: List[Dict[str, Any]] = getattr(col, 'advanced_rules', None) or []
    has_else: Optional[str] = getattr(col, 'advanced_else', None)
    if not rules and (has_else is None or has_else == ""):
        return current_value
    ref_values: Dict[str, str] = {k: safe_str(v) for k, v in row_map.items()}
    for r in rules:
        ref: Optional[str] = r.get("ref") or r.get("ref_target")
        op: str = (r.get("op") or "equals").lower()
        match: str = r.get("match", "")
        set_val: Any = r.get("set")
        ref_val: str = ref_values.get(ref, "") if ref else ""
        cond: bool = False
        if op == "equals":
            cond = (ref_val == match)
        elif op == "not_equals":
            cond = (ref_val != match)
        elif op == "contains":
            cond = (match in ref_val)
        elif op in ("in", "not_in"):
            tokens: List[str] = [t.strip() for t in safe_str(match).split(",")]
            in_set: bool = ref_val in tokens
            cond = in_set if op == "in" else (not in_set)
        if cond:
            return set_val if set_val is not None else current_value
    if has_else is not None and has_else != "":
        return has_else
    return current_value

def generate_preview_data(spec: MappingSpec, max_rows_per_sheet: int = 1000) -> Dict[str, Dict[str, Any]]:
    if not spec.source_path:
        raise ValueError("spec.source_path is not set")
    src_wb: Workbook = load_workbook(spec.source_path, read_only=True, data_only=True)
    preview: Dict[str, Dict[str, Any]] = {}
    src_headers_by_sheet: Dict[str, List[str]] = read_workbook_headers(spec.source_path)

    for sm in spec.sheets:
        headers: List[str] = list(sm.target_headers)
        rows: List[List[Any]] = []
        truncated: bool = False

        if not sm.source_sheet or sm.source_sheet not in src_wb.sheetnames:
            preview[sm.target_sheet] = {"headers": headers, "rows": rows, "truncated": False}
            continue

        ws_src: Worksheet = src_wb[sm.source_sheet]
        src_headers: List[str] = src_headers_by_sheet.get(sm.source_sheet, [])
        src_index: Dict[str, int] = {h: i for i, h in enumerate(src_headers)}

        col_map: List[Dict[str, Any]] = []
        for col in sm.columns:
            entry: Dict[str, Any] = {
                "target": col.target,
                "source_idx": src_index.get(col.source) if col.source in src_index else None,
                "default": col.default,
                "transforms": col.transforms,
                "find_replace": col.find_replace,
                "data_type": getattr(col, "data_type", None),
                "number_format": getattr(col, "number_format", ""),
            }
            col_map.append(entry)

        count: int = 0
        for row in ws_src.iter_rows(min_row=2, values_only=True):
            out_row: List[Any] = []
            for entry in col_map:
                val: Any = None
                if entry["source_idx"] is not None and entry["source_idx"] < len(row):
                    val = row[entry["source_idx"]]
                val = apply_transforms(val, entry["transforms"])
                val = replace_values(val, spec.global_find_replace)
                val = replace_values(val, entry["find_replace"])
                if is_blank(val) and entry["default"] not in (None, ""):
                    val = entry["default"]
                val = coerce_value(val, entry.get("data_type"))
                out_row.append(val)

            row_map: Dict[str, Any] = {h: v for h, v in zip(sm.target_headers, out_row)}
            for idx, col in enumerate(sm.columns):
                out_row[idx] = apply_advanced_to_cell(col, row_map, out_row[idx])

            if sm.drop_if_all_blank and all(is_blank(v) for v in out_row):
                continue

            rows.append(out_row)
            count += 1
            if count >= max_rows_per_sheet:
                truncated = True
                break

        preview[sm.target_sheet] = {"headers": headers, "rows": rows, "truncated": truncated}

    src_wb.close()
    return preview

def apply_template(spec: MappingSpec, output_path: str) -> None:
    if not spec.source_path:
        raise ValueError("spec.source_path is not set")

    src_wb: Workbook = load_workbook(spec.source_path, read_only=True, data_only=True)

    out_wb: Workbook = Workbook()
    if out_wb.active and len(out_wb.worksheets) == 1 and out_wb.active.title == "Sheet":
        out_wb.remove(out_wb.active)

    src_headers_by_sheet: Dict[str, List[str]] = read_workbook_headers(spec.source_path)

    for sm in spec.sheets:
        ws_out: Worksheet = out_wb.create_sheet(title=sm.target_sheet)
        for c_idx, header in enumerate(sm.target_headers, start=1):
            ws_out.cell(row=1, column=c_idx, value=header)

        if not sm.source_sheet or sm.source_sheet not in src_wb.sheetnames:
            continue

        ws_src: Worksheet = src_wb[sm.source_sheet]
        src_headers: List[str] = src_headers_by_sheet.get(sm.source_sheet, [])
        src_index: Dict[str, int] = {h: i for i, h in enumerate(src_headers)}

        col_map: List[Dict[str, Any]] = []
        for col in sm.columns:
            entry: Dict[str, Any] = {
                "target": col.target,
                "source_idx": src_index.get(col.source) if col.source in src_index else None,
                "default": col.default,
                "transforms": col.transforms,
                "find_replace": col.find_replace,
                "data_type": getattr(col, "data_type", None),
                "number_format": getattr(col, "number_format", ""),
            }
            col_map.append(entry)

        out_row_idx: int = 2
        for row in ws_src.iter_rows(min_row=2, values_only=True):
            out_row: List[Any] = []
            for entry in col_map:
                val: Any = None
                if entry["source_idx"] is not None and entry["source_idx"] < len(row):
                    val = row[entry["source_idx"]]
                val = apply_transforms(val, entry["transforms"])
                val = replace_values(val, spec.global_find_replace)
                val = replace_values(val, entry["find_replace"])
                if is_blank(val) and entry["default"] not in (None, ""):
                    val = entry["default"]
                val = coerce_value(val, entry.get("data_type"))
                out_row.append(val)

            row_map: Dict[str, Any] = {h: v for h, v in zip(sm.target_headers, out_row)}
            for idx, col in enumerate(sm.columns):
                out_row[idx] = apply_advanced_to_cell(col, row_map, out_row[idx])

            if sm.drop_if_all_blank and all(is_blank(v) for v in out_row):
                continue

            for c_idx, (v, entry) in enumerate(zip(out_row, col_map), start=1):
                cell: Union[Cell, MergedCell] = ws_out.cell(row=out_row_idx, column=c_idx, value=v)
                nf: str = entry.get("number_format", "")
                if nf and isinstance(cell, Cell):
                    try:
                        cell.number_format = nf
                    except Exception:
                        pass
            out_row_idx += 1

    out_wb.save(output_path)
    src_wb.close()