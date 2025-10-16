import difflib
from typing import Dict, List, Any, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

def is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False

def safe_str(value: Any) -> str:
    try:
        return "" if value is None else str(value)
    except Exception:
        return ""

def read_workbook_headers(path: str) -> Dict[str, List[str]]:
    wb: Workbook = load_workbook(path, read_only=True, data_only=True)
    headers_by_sheet: Dict[str, List[str]] = {}
    for ws in wb.worksheets:
        headers: List[str] = []
        first_row: Optional[Tuple[Any, ...]] = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if first_row:
            headers = [safe_str(h).strip() for h in first_row]
            while headers and headers[-1] == "":
                headers.pop()
        headers_by_sheet[ws.title] = headers
    wb.close()
    return headers_by_sheet

def suggest_header_mapping(target_headers: List[str], source_headers: List[str]) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    source_lower: Dict[str, str] = {h.lower(): h for h in source_headers}
    for th in target_headers:
        if th in source_headers:
            mapping[th] = th
            continue
        lower: str = th.lower()
        if lower in source_lower:
            mapping[th] = source_lower[lower]
            continue
        candidates: List[str] = difflib.get_close_matches(th, source_headers, n=1, cutoff=0.82)
        mapping[th] = candidates[0] if candidates else ""
    return mapping